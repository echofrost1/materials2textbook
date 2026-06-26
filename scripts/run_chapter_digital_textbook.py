#!/usr/bin/env python
"""Run chapter-scoped digital textbook generation.

This is the role-B entry point from
docs/two_person_textbook_project_task_allocation.md.  It consumes the chapter
evidence pack produced by role A, generates a single chapter, and writes gap
feedback for material issues instead of silently expanding thin evidence.
"""

from __future__ import annotations

import argparse
import json
import math
import re
import subprocess
import sys
from datetime import datetime
from pathlib import Path

from material_paths import default_raw_root, default_work_root
from typing import Any

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_MATERIAL_ROOT = default_work_root()

CHAPTERS: dict[str, str] = {
    "tig_welding": "钨极氩弧焊",
    "welding_equipment_safety": "焊接设备与安全",
    "welding_basic_operation": "焊接基本操作",
    "shielded_metal_arc_welding": "焊条电弧焊",
    "gas_welding_and_cutting": "气焊与气割",
    "welding_quality_inspection": "焊接质量检验",
    "textbook_reference": "教材参考资料",
}

WRITABLE_READINESS = {"ready", "partial_ready"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run one chapter through the digital textbook workflow.")
    parser.add_argument("--material-root", type=Path, default=DEFAULT_MATERIAL_ROOT)
    parser.add_argument(
        "--chapter-code",
        default="tig_welding",
        help="Stable chapter code. Defaults to tig_welding to avoid Chinese argument encoding issues.",
    )
    parser.add_argument("--chapter", default="", help="Optional chapter title override.")
    parser.add_argument("--title", default=None)
    parser.add_argument("--use-llm", action="store_true")
    parser.add_argument("--max-input-tokens", type=int, default=120000)
    parser.add_argument("--max-chunks-per-knowledge-point", type=int, default=12)
    parser.add_argument("--max-video-records", type=int, default=0)
    parser.add_argument("--max-document-records", type=int, default=0)
    parser.add_argument("--review-rounds", type=int, default=1)
    parser.add_argument("--skip-build-pack", action="store_true")
    parser.add_argument("--knowledge-points", default="", help="Optional semicolon-separated readiness target points.")
    parser.add_argument(
        "--include-not-ready",
        action="store_true",
        help="Allow not_ready evidence into draft generation. Default keeps it only in the gap log.",
    )
    parser.add_argument(
        "--preview-only",
        action="store_true",
        help="Generate local preview without packaging zip. This is the default when --package-offline is absent.",
    )
    parser.add_argument(
        "--package-offline",
        action="store_true",
        help="Copy required media into digital_book/assets and validate digital_book.zip for offline sharing.",
    )
    parser.add_argument(
        "--copy-media-assets",
        action="store_true",
        help="Copy media into the preview digital_book directory. Implied by --package-offline.",
    )
    parser.add_argument("--student-package-max-mb", type=float, default=2048.0)
    parser.add_argument("--student-package-max-asset-files", type=int, default=0)
    parser.add_argument("--student-package-skip-static-smoke", action="store_true")
    parser.add_argument("--student-package-asset-fallback-zip", type=Path, default=None)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.preview_only and args.package_offline:
        raise SystemExit("--preview-only and --package-offline cannot be used together.")

    material_root = args.material_root.resolve()
    chapter_code, chapter_title = resolve_chapter(args.chapter_code, args.chapter)
    chapter_root = material_root / "05_final_deliverables" / "chapter_work" / chapter_code
    output_dir = chapter_root / "agent_workflow"
    writer_input_root = chapter_root / "writer_inputs"
    package_output = chapter_root / "digital_book.zip"
    cache_path = output_dir / "llm_cache.json"

    if not args.skip_build_pack:
        build_command = [
            sys.executable,
            str(ROOT / "scripts" / "build_chapter_evidence_pack.py"),
            "--material-root",
            str(material_root),
            "--chapter",
            chapter_title,
            "--output-root",
            str(chapter_root),
        ]
        if args.knowledge_points.strip():
            build_command.extend(["--knowledge-points", args.knowledge_points])
        run(build_command)

    pack_path = chapter_root / "chapter_evidence_pack.jsonl"
    readiness_path = chapter_root / "chapter_readiness_report.xlsx"
    if not pack_path.exists():
        raise SystemExit(f"Missing chapter evidence pack: {pack_path}")
    if not readiness_path.exists():
        raise SystemExit(f"Missing chapter readiness report: {readiness_path}")

    readiness_rows = load_readiness(readiness_path)
    gap_log_rows = build_gap_log(chapter_title, readiness_rows)
    write_gap_log(chapter_root, gap_log_rows)

    video_segments, ppt_assets, document_segments, selected_count = build_writer_inputs_from_pack(
        pack_path=pack_path,
        readiness_rows=readiness_rows,
        output_root=writer_input_root,
        include_not_ready=args.include_not_ready,
    )
    if selected_count == 0:
        raise SystemExit(
            "No ready or partial_ready evidence found for writing. "
            "Check chapter_readiness_report.xlsx or rerun with --include-not-ready for diagnostics."
        )

    command = [
        sys.executable,
        str(ROOT / "scripts" / "run_full_digital_textbook.py"),
        "--material-root",
        str(material_root),
        "--title",
        args.title or f"{chapter_title}数字教材样章",
        "--segments",
        str(video_segments),
        "--ppt-assets",
        str(ppt_assets),
        "--document-segments",
        str(document_segments),
        "--output-dir",
        str(output_dir),
        "--max-video-records",
        str(args.max_video_records),
        "--max-document-records",
        str(args.max_document_records),
        "--max-input-tokens",
        str(args.max_input_tokens),
        "--max-tokens-per-evidence-chunk",
        "1200",
        "--summarize-over-budget",
        "--summary-token-reserve-ratio",
        "0.25",
        "--max-tokens-per-summary-chunk",
        "700",
        "--max-summary-source-chunks",
        "8",
        "--max-chunks-per-knowledge-point",
        str(args.max_chunks_per_knowledge_point),
        "--review-rounds",
        str(args.review_rounds),
        "--llm-cache-path",
        str(cache_path),
    ]
    if args.use_llm:
        command.append("--use-llm")
    if args.copy_media_assets or args.package_offline:
        command.append("--copy-media-assets")
    if args.package_offline:
        command.extend(
            [
                "--student-package-output",
                str(package_output),
                "--student-package-max-mb",
                str(args.student_package_max_mb),
                "--student-package-max-asset-files",
                str(args.student_package_max_asset_files),
            ]
        )
        if args.student_package_asset_fallback_zip:
            command.extend(["--student-package-asset-fallback-zip", str(args.student_package_asset_fallback_zip)])
        if args.student_package_skip_static_smoke:
            command.append("--student-package-skip-static-smoke")

    run(command)

    report = build_generation_report(
        chapter_code=chapter_code,
        chapter_title=chapter_title,
        chapter_root=chapter_root,
        output_dir=output_dir,
        writer_input_root=writer_input_root,
        gap_log_rows=gap_log_rows,
        use_llm_requested=args.use_llm,
        output_mode="package_offline" if args.package_offline else "preview_only",
    )
    report_json_path = chapter_root / "chapter_generation_report.json"
    report_markdown_path = chapter_root / "chapter_generation_report.md"
    write_json(report_json_path, report)
    write_text(report_markdown_path, render_generation_report_markdown(report))

    print("Chapter digital textbook generated:")
    print(f"- chapter_code: {chapter_code}")
    print(f"- chapter_title: {chapter_title}")
    print(f"- chapter_root: {chapter_root}")
    print(f"- readiness_report: {readiness_path}")
    print(f"- gap_log: {chapter_root / 'chapter_evidence_gap_log.xlsx'}")
    print(f"- writer_inputs: {writer_input_root}")
    print(f"- final_markdown: {output_dir / 'textbook_final.md'}")
    print(f"- generation_report: {report_markdown_path}")
    print(f"- digital_book_index: {chapter_root / 'digital_book' / 'index.html'}")
    print(f"- output_mode: {'package_offline' if args.package_offline else 'preview_only'}")
    if args.package_offline:
        print(f"- package: {package_output}")
    return 0


def resolve_chapter(chapter_code: str, chapter_override: str = "") -> tuple[str, str]:
    code = chapter_code.strip()
    if not code:
        code = "tig_welding"
    if chapter_override.strip():
        title = chapter_override.strip()
        return code or safe_slug(title), title
    title = CHAPTERS.get(code)
    if not title:
        known = ", ".join(sorted(CHAPTERS))
        raise SystemExit(f"Unknown --chapter-code {code!r}. Known codes: {known}. Or pass --chapter.")
    return code, title


def load_readiness(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [text(value) for value in rows[0]]
    return [
        normalize_row({headers[index]: value for index, value in enumerate(row) if index < len(headers)})
        for row in rows[1:]
    ]


def build_gap_log(chapter_title: str, readiness_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    now = datetime.now().isoformat(timespec="seconds")
    for row in readiness_rows:
        status = text(row.get("readiness_status")).lower()
        if status == "ready":
            continue
        point = text(row.get("knowledge_point")) or "未命名知识点"
        gap_type = "not_ready" if status == "not_ready" else "partial_ready"
        rows.append(
            {
                "chapter": chapter_title,
                "knowledge_point": point,
                "gap_type": gap_type,
                "gap_description": text(row.get("readiness_reason")) or "evidence is not ready for expanded writing",
                "video_segment_count": row.get("video_segment_count", 0),
                "primary_video_count": row.get("primary_video_count", 0),
                "ppt_page_count": row.get("ppt_page_count", 0),
                "reference_text_count": row.get("reference_text_count", 0),
                "usable_evidence_count": int_value(row.get("pass_count")) + int_value(row.get("weak_count")),
                "found_by": "role_b_chapter_runner",
                "assigned_to": "role_a_material_library",
                "status": "open" if status == "not_ready" else "monitor",
                "resolution": "",
                "updated_at": now,
            }
        )
    return rows


def write_gap_log(chapter_root: Path, rows: list[dict[str, Any]]) -> None:
    chapter_root.mkdir(parents=True, exist_ok=True)
    xlsx_path = chapter_root / "chapter_evidence_gap_log.xlsx"
    jsonl_path = chapter_root / "chapter_evidence_gap_log.jsonl"
    write_xlsx(xlsx_path, rows)
    write_jsonl(jsonl_path, rows)


def build_writer_inputs_from_pack(
    *,
    pack_path: Path,
    readiness_rows: list[dict[str, Any]],
    output_root: Path,
    include_not_ready: bool = False,
) -> tuple[Path, Path, Path, int]:
    allowed_points = writable_points(readiness_rows, include_not_ready=include_not_ready)
    pack_rows = [row for row in read_jsonl(pack_path) if text(row.get("knowledge_point")) in allowed_points]
    pack_rows = [
        row
        for row in pack_rows
        if text(row.get("quality_gate_status")).lower() in {"pass", "weak"}
        and text(row.get("evidence_role")).lower() != "hold"
    ]
    video_rows: list[dict[str, Any]] = []
    ppt_rows: list[dict[str, Any]] = []
    document_rows: list[dict[str, Any]] = []
    for row in pack_rows:
        source_type = text(row.get("source_type")).lower()
        if source_type in {"video", "video_segment", "audio", "audio_segment"}:
            video_rows.append(pack_row_to_video_segment(row))
        elif source_type in {"ppt", "ppt_slide", "slide"}:
            ppt_rows.append(pack_row_to_document_segment(row, source_type="ppt_slide"))
        else:
            document_rows.append(pack_row_to_document_segment(row, source_type=source_type or "document_segment"))

    output_root.mkdir(parents=True, exist_ok=True)
    video_path = output_root / "chapter_video_segments.from_pack.jsonl"
    ppt_path = output_root / "chapter_ppt_assets.from_pack.jsonl"
    document_path = output_root / "chapter_document_segments.from_pack.jsonl"
    write_jsonl(video_path, video_rows)
    write_jsonl(ppt_path, ppt_rows)
    write_jsonl(document_path, document_rows)
    return video_path, ppt_path, document_path, len(pack_rows)


def writable_points(readiness_rows: list[dict[str, Any]], *, include_not_ready: bool = False) -> set[str]:
    allowed = set()
    for row in readiness_rows:
        point = text(row.get("knowledge_point"))
        status = text(row.get("readiness_status")).lower()
        if not point:
            continue
        if include_not_ready or status in WRITABLE_READINESS:
            allowed.add(point)
    return allowed


def pack_row_to_video_segment(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "clip_id": text(row.get("chunk_id")),
        "source_asset_id": text(row.get("asset_id")),
        "knowledge_point": text(row.get("knowledge_point")),
        "evidence_text": text(row.get("evidence_text")),
        "transcript_text": text(row.get("evidence_text")),
        "clip_summary": text(row.get("summary")) or text(row.get("title")),
        "source_video": text(row.get("source_path")),
        "original_path": text(row.get("source_path")),
        "start_time": text(row.get("start_time")),
        "end_time": text(row.get("end_time")),
        "keyframe_paths": text(row.get("image_paths")),
        "review_status": text(row.get("review_status")),
        "usefulness_score": float_value(row.get("teaching_value")),
        "quality_score": float_value(row.get("confidence")),
        "recommended_chapter": text(row.get("chapter")),
        "material_block": text(row.get("chapter")),
        "tags": text(row.get("tags")),
    }


def pack_row_to_document_segment(row: dict[str, Any], *, source_type: str) -> dict[str, Any]:
    segment_id = text(row.get("chunk_id"))
    payload = {
        "segment_id": segment_id,
        "chunk_id": segment_id,
        "asset_id": text(row.get("asset_id")),
        "knowledge_point": text(row.get("knowledge_point")),
        "heading": text(row.get("title")),
        "title": text(row.get("title")),
        "summary": text(row.get("summary")),
        "evidence_text": text(row.get("evidence_text")),
        "source_type": source_type,
        "document_type": source_type,
        "document_path": text(row.get("source_path")),
        "original_path": text(row.get("source_path")),
        "page": row.get("page_or_slide", ""),
        "slide_index": row.get("page_or_slide", ""),
        "image_paths": text(row.get("image_paths")),
        "review_status": text(row.get("review_status")),
        "teaching_value": float_value(row.get("teaching_value")),
        "relevance_score": float_value(row.get("relevance") or row.get("relevance_score")),
        "confidence": float_value(row.get("confidence")),
        "recommended_chapter": text(row.get("chapter")),
        "chapter": text(row.get("chapter")),
        "material_block": text(row.get("chapter")),
        "tags": text(row.get("tags")),
    }
    if source_type == "ppt_slide":
        payload["ppt_asset_id"] = segment_id
        payload["slide_title"] = text(row.get("title"))
        payload["slide_text"] = text(row.get("evidence_text"))
    return payload


def run(command: list[str]) -> None:
    printable = " ".join(f'"{part}"' if " " in part else part for part in command)
    print(f"[chapter-runner] {printable}", flush=True)
    subprocess.run(command, cwd=ROOT, check=True)


def build_generation_report(
    *,
    chapter_code: str,
    chapter_title: str,
    chapter_root: Path,
    output_dir: Path,
    writer_input_root: Path,
    gap_log_rows: list[dict[str, Any]],
    use_llm_requested: bool,
    output_mode: str,
) -> dict[str, Any]:
    final_markdown_path = output_dir / "textbook_final.md"
    manifest_path = output_dir / "artifact_manifest.json"
    cache_path = output_dir / "llm_cache.json"
    final_markdown = final_markdown_path.read_text(encoding="utf-8") if final_markdown_path.exists() else ""
    manifest = read_json(manifest_path)
    summary = manifest.get("summary", {}) if isinstance(manifest, dict) else {}

    report = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "chapter_code": chapter_code,
        "chapter_title": chapter_title,
        "chapter_root": str(chapter_root),
        "output_mode": output_mode,
        "use_llm_requested": use_llm_requested,
        "writer_generation_mode": summary.get("writer_generation_mode", ""),
        "writer_generation_warning": summary.get("writer_generation_warning", ""),
        "llm_cache_exists": cache_path.exists(),
        "llm_cache_entries": count_llm_cache_entries(cache_path),
        "final_markdown_chars": len(final_markdown),
        "contains_learning_path_heading": "### 学习路径" in final_markdown,
        "contains_raw_evidence_list": "- 证据 `" in final_markdown,
        "gap_log_rows": len(gap_log_rows),
        "gap_log_by_type": count_by_key(gap_log_rows, "gap_type"),
        "writer_input_counts": {
            "video_segments": count_jsonl(writer_input_root / "chapter_video_segments.from_pack.jsonl"),
            "ppt_assets": count_jsonl(writer_input_root / "chapter_ppt_assets.from_pack.jsonl"),
            "document_segments": count_jsonl(writer_input_root / "chapter_document_segments.from_pack.jsonl"),
        },
        "workflow_summary": {
            key: summary.get(key)
            for key in [
                "evidence_chunks",
                "skipped_chunks",
                "token_budget_enabled",
                "token_budget_max_input_tokens",
                "token_budget_original_estimated_tokens",
                "token_budget_kept_estimated_tokens",
                "token_budget_kept_source_chunks",
                "token_budget_truncated_chunks",
                "token_budget_dropped_chunks",
                "token_budget_summary_chunks",
                "token_budget_summarized_source_chunks",
                "token_budget_uncovered_dropped_chunks",
                "chapters",
                "knowledge_points",
                "fact_issue_count",
                "pedagogy_issue_count",
                "review_rounds_completed",
                "citation_coverage_rate",
                "paragraph_support_rate",
                "claim_support_rate",
                "overall_quality_score",
            ]
        },
    }
    return report


def render_generation_report_markdown(report: dict[str, Any]) -> str:
    workflow = report.get("workflow_summary", {})
    inputs = report.get("writer_input_counts", {})
    gaps = report.get("gap_log_by_type", {})
    lines = [
        f"# {report.get('chapter_title', '')}单章生成报告",
        "",
        "## 生成状态",
        f"- 章节代码：{report.get('chapter_code', '')}",
        f"- 输出模式：{report.get('output_mode', '')}",
        f"- 请求 LLM：{yes_no(report.get('use_llm_requested'))}",
        f"- 写作模式：{report.get('writer_generation_mode') or 'unknown'}",
        f"- LLM 缓存：{yes_no(report.get('llm_cache_exists'))}，条目 {report.get('llm_cache_entries', 0)}",
        f"- 最终正文字符数：{report.get('final_markdown_chars', 0)}",
        "",
        "## 内容风险",
        f"- 是否残留学习路径标题：{yes_no(report.get('contains_learning_path_heading'))}",
        f"- 是否残留原始证据列表：{yes_no(report.get('contains_raw_evidence_list'))}",
        f"- 证据缺口行数：{report.get('gap_log_rows', 0)}，partial_ready={gaps.get('partial_ready', 0)}，not_ready={gaps.get('not_ready', 0)}",
        "",
        "## 输入规模",
        f"- 视频片段：{inputs.get('video_segments', 0)}",
        f"- PPT/图片素材：{inputs.get('ppt_assets', 0)}",
        f"- 文档片段：{inputs.get('document_segments', 0)}",
        "",
        "## 质量指标",
        f"- 证据块：{workflow.get('evidence_chunks', 0)}",
        f"- 知识点：{workflow.get('knowledge_points', 0)}",
        f"- citation_coverage_rate：{workflow.get('citation_coverage_rate', '')}",
        f"- paragraph_support_rate：{workflow.get('paragraph_support_rate', '')}",
        f"- claim_support_rate：{workflow.get('claim_support_rate', '')}",
        f"- fact_issue_count：{workflow.get('fact_issue_count', '')}",
        f"- overall_quality_score：{workflow.get('overall_quality_score', '')}",
    ]
    warning = report.get("writer_generation_warning")
    if warning:
        lines.extend(["", "## 写作警告", str(warning)])
    return "\n".join(lines).rstrip() + "\n"


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open(encoding="utf-8") as handle:
        for line in handle:
            if line.strip():
                rows.append(json.loads(line))
    return rows


def read_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def count_jsonl(path: Path) -> int:
    if not path.exists():
        return 0
    with path.open(encoding="utf-8") as handle:
        return sum(1 for line in handle if line.strip())


def count_llm_cache_entries(path: Path) -> int:
    payload = read_json(path)
    if isinstance(payload, dict):
        return len(payload)
    return 0


def count_by_key(rows: list[dict[str, Any]], key: str) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in rows:
        value = text(row.get(key)) or "unknown"
        counts[value] = counts.get(value, 0) + 1
    return counts


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")


def normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    return {key: clean_cell(value) for key, value in row.items()}


def clean_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    if isinstance(value, (list, dict, tuple, set)):
        return value
    return value


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    if isinstance(value, (list, tuple, set)):
        return ";".join(text(item) for item in value if text(item))
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return str(value).strip()


def yes_no(value: Any) -> str:
    return "是" if bool(value) else "否"


def int_value(value: Any) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def float_value(value: Any) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def write_xlsx(path: Path, rows: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    if rows:
        headers = list(rows[0].keys())
        sheet.append(headers)
        for row in rows:
            sheet.append([excel_cell(row.get(header, "")) for header in headers])
    else:
        sheet.append(
            [
                "chapter",
                "knowledge_point",
                "gap_type",
                "gap_description",
                "found_by",
                "assigned_to",
                "status",
                "resolution",
                "updated_at",
            ]
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def excel_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return text(value)


def safe_slug(value: str) -> str:
    slug = re.sub(r"[^0-9A-Za-z_\u4e00-\u9fff]+", "_", value).strip("_")
    return slug or "chapter"


if __name__ == "__main__":
    raise SystemExit(main())
