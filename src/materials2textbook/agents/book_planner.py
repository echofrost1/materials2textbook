from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook

from materials2textbook.schemas import (
    BookChapterPlan,
    BookPlan,
    BookSectionPlan,
    ChapterPlan,
    EvidenceChunk,
    KnowledgePoint,
    ReviewIssue,
)


DEFAULT_CHAPTER_TOKEN_BUDGET = 12000
DEFAULT_CHAPTER_VIDEO_BUDGET = 3
DEFAULT_CHAPTER_DOCUMENT_BUDGET = 20
DEFAULT_MAX_KNOWLEDGE_POINTS = 8
DEFAULT_WELDING_CHAPTER_ORDER = [
    "焊接设备与安全",
    "焊接基本操作",
    "焊条电弧焊",
    "钨极氩弧焊",
    "二氧化碳气体保护焊",
    "气焊与气割",
    "焊接质量检验",
    "综合训练与考核",
    "机械制图-投影法",
    "工程材料-性能测试",
    "焊接鉴定与试题",
    "教材参考资料",
]


@dataclass
class ManifestMaterial:
    material_id: str
    title: str
    chapter: str = ""
    section: str = ""
    knowledge_point: str = ""
    source_type: str = ""
    path: str = ""
    notes: str = ""
    chapter_code: str = ""
    section_code: str = ""
    confidence: float = 1.0
    active_for_index: bool = True
    needs_confirmation: bool = False


class BookPlannerAgent:
    """Plan a whole textbook before chapter-level generation."""

    def run(
        self,
        *,
        title: str,
        chunks: list[EvidenceChunk],
        manifest_xlsx: Path | None = None,
        max_chapters: int = 0,
        max_knowledge_points_per_chapter: int = DEFAULT_MAX_KNOWLEDGE_POINTS,
        chapter_token_budget: int = DEFAULT_CHAPTER_TOKEN_BUDGET,
    ) -> BookPlan:
        manifest_rows = read_manifest_xlsx(manifest_xlsx) if manifest_xlsx else []
        manifest_by_id = _best_manifest_by_id(manifest_rows)
        curriculum_order = build_curriculum_order(manifest_rows, chunks)
        curriculum_index = {title: index for index, title in enumerate(curriculum_order, start=1)}
        chunk_ids = {chunk.chunk_id for chunk in chunks}
        primary_owner: dict[str, str] = {}
        chapters: dict[str, list[EvidenceChunk]] = defaultdict(list)
        reference_by_chapter: dict[str, list[str]] = defaultdict(list)

        for chunk in chunks:
            manifest = _manifest_for_chunk(chunk, manifest_by_id)
            chapter_title = _first_text(
                manifest.chapter if manifest else "",
                chunk.metadata.get("chapter"),
                chunk.metadata.get("chapter_title"),
                chunk.metadata.get("target_chapter"),
                chunk.recommended_chapter,
                "待规划章节",
            )
            owner = primary_owner.get(chunk.chunk_id)
            if owner and owner != chapter_title:
                reference_by_chapter[chapter_title].append(chunk.chunk_id)
                continue
            primary_owner[chunk.chunk_id] = chapter_title
            chapters[chapter_title].append(chunk)

        ordered_chapters = sorted(
            chapters.items(),
            key=lambda item: (_chapter_order(item[0], item[1], manifest_rows, curriculum_index), item[0]),
        )
        if max_chapters > 0:
            ordered_chapters = ordered_chapters[:max_chapters]

        book_chapters: list[BookChapterPlan] = []
        for chapter_index, (chapter_title, chapter_chunks) in enumerate(ordered_chapters, start=1):
            chapter_id = f"chapter_{chapter_index:02d}"
            sections = _build_sections(
                chapter_id=chapter_id,
                chapter_no=chapter_index,
                chunks=chapter_chunks,
                manifest_by_id=manifest_by_id,
                max_knowledge_points=max_knowledge_points_per_chapter,
            )
            primary_ids = _budget_primary_materials(chapter_chunks)
            chapter_reference_ids = _dedupe(
                [
                    chunk.chunk_id
                    for chunk in chapter_chunks
                    if chunk.chunk_id not in primary_ids
                ]
                + reference_by_chapter.get(chapter_title, [])
            )
            book_chapters.append(
                BookChapterPlan(
                    chapter_id=chapter_id,
                    chapter_no=chapter_index,
                    title=chapter_title,
                    learning_goals=[
                        f"理解{chapter_title}的核心概念、材料来源和学习任务。",
                        f"能够结合教材资源说明{chapter_title}的关键知识点。",
                        "能够通过视频、案例和练习完成迁移判断。",
                    ],
                    sections=sections,
                    primary_material_ids=primary_ids,
                    reference_material_ids=chapter_reference_ids,
                    token_budget=chapter_token_budget,
                    video_budget=DEFAULT_CHAPTER_VIDEO_BUDGET,
                    document_budget=DEFAULT_CHAPTER_DOCUMENT_BUDGET,
                )
            )

        return BookPlan(
            book_id=_slugify(title),
            title=title,
            planning_strategy="manifest_xlsx_first",
            chapters=book_chapters,
            material_stats={
                "evidence_chunks": len(chunks),
                "manifest_rows": len(manifest_rows),
                "manifest_matched_chunks": sum(1 for chunk in chunks if _manifest_for_chunk(chunk, manifest_by_id)),
                "unassigned_chunks": len(chunk_ids - set(primary_owner)),
            },
            budget={
                "chapter_token_budget": chapter_token_budget,
                "chapter_video_budget": DEFAULT_CHAPTER_VIDEO_BUDGET,
                "chapter_document_budget": DEFAULT_CHAPTER_DOCUMENT_BUDGET,
                "max_knowledge_points_per_chapter": max_knowledge_points_per_chapter,
            },
            metadata={
                "manifest_xlsx": str(manifest_xlsx or ""),
                "curriculum_order": curriculum_order,
                "curriculum_order_source": "manifest_plus_default_welding_rules",
            },
        )


def read_manifest_xlsx(path: Path | None) -> list[ManifestMaterial]:
    if not path:
        return []
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Manifest xlsx not found: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[ManifestMaterial] = []
    for sheet in workbook.worksheets:
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            continue
        headers = [str(value or "").strip() for value in values[0]]
        for raw_values in values[1:]:
            normalized = {
                _normalize_header(header): str(value or "").strip()
                for header, value in zip(headers, raw_values, strict=False)
            }
            material_id = _first_text(
                _field(normalized, "asset_id", "material_id", "素材id", "资源id", "编号", "id"),
                _field(normalized, "chunk_id", "片段id"),
                _field(normalized, "文件名", "filename", "file_name"),
            )
            title = _first_text(
                _field(normalized, "knowledge_point", "知识点", "知识点名称"),
                _field(normalized, "title", "标题", "素材名称", "资源名称"),
                material_id,
            )
            if not title and not material_id:
                continue
            rows.append(
                ManifestMaterial(
                    material_id=material_id,
                    title=title,
                    chapter=_field(
                        normalized,
                        "chapter",
                        "chapter_title",
                        "recommended_chapter",
                        "target_chapter",
                        "material_block_cn",
                        "material_block",
                        "章节",
                        "章",
                        "章标题",
                    ),
                    section=_field(
                        normalized,
                        "section",
                        "section_title",
                        "target_section",
                        "knowledge_point_cn",
                        "knowledge_point",
                        "节",
                        "小节",
                        "节标题",
                    ),
                    knowledge_point=_field(normalized, "knowledge_point", "knowledge_point_cn", "知识点", "知识点名称"),
                    source_type=_field(normalized, "source_type", "asset_type", "file_type", "素材类型", "资源类型", "类型"),
                    path=_field(normalized, "path", "文件路径", "路径", "original_path"),
                    notes=_field(normalized, "notes", "备注", "说明"),
                    chapter_code=_field(normalized, "chapter_code", "material_block_code", "章编码", "板块编码"),
                    section_code=_field(normalized, "section_code", "knowledge_point_code", "节编码", "知识点编码"),
                    confidence=_parse_float(_field(normalized, "confidence", "classification_confidence", "置信度"), 1.0),
                    active_for_index=_parse_bool(_field(normalized, "active_for_index", "active_for_processing", "是否启用"), default=True),
                    needs_confirmation=_parse_bool(_field(normalized, "needs_confirmation", "需要确认"), default=False),
                )
            )
    return rows


def build_curriculum_order(manifest_rows: list[ManifestMaterial], chunks: list[EvidenceChunk]) -> list[str]:
    """Build a deterministic course order from manifest data plus welding defaults."""

    counts: Counter[str] = Counter()
    code_by_chapter: dict[str, str] = {}
    for row in manifest_rows:
        chapter = row.chapter.strip()
        if not chapter:
            continue
        counts[chapter] += 1
        if row.chapter_code and chapter not in code_by_chapter:
            code_by_chapter[chapter] = row.chapter_code
    for chunk in chunks:
        chapter = _first_text(
            chunk.metadata.get("chapter"),
            chunk.metadata.get("chapter_title"),
            chunk.metadata.get("target_chapter"),
            chunk.material_block,
            chunk.recommended_chapter,
        )
        if chapter:
            counts[chapter] += 1
            if chunk.material_block_code and chapter not in code_by_chapter:
                code_by_chapter[chapter] = chunk.material_block_code

    result: list[str] = []
    for title in DEFAULT_WELDING_CHAPTER_ORDER:
        match = _find_matching_chapter(title, counts)
        if match and match not in result:
            result.append(match)

    remaining = [title for title in counts if title not in result]
    remaining.sort(key=lambda title: (_code_sort_key(code_by_chapter.get(title, "")), -counts[title], title))
    result.extend(remaining)
    return result


def render_curriculum_order_yaml(book_plan: BookPlan) -> str:
    order = list(book_plan.metadata.get("curriculum_order") or [chapter.title for chapter in book_plan.chapters])
    lines = [
        "# Auto-generated by materials2textbook.",
        "# It is deterministic and can be regenerated from the manifest; manual review is optional.",
        f"title: {book_plan.title}",
        f"source: {book_plan.metadata.get('curriculum_order_source', 'auto')}",
        "chapter_order:",
    ]
    for index, title in enumerate(order, start=1):
        chapter = next((item for item in book_plan.chapters if item.title == title), None)
        if chapter:
            lines.append(f"  - order: {index}")
            lines.append(f"    title: {title}")
            lines.append(f"    sections: {len(chapter.sections)}")
            lines.append(f"    primary_materials: {len(chapter.primary_material_ids)}")
        else:
            lines.append(f"  - order: {index}")
            lines.append(f"    title: {title}")
            lines.append("    sections: 0")
            lines.append("    primary_materials: 0")
    return "\n".join(lines) + "\n"


def book_plan_to_chapter_plans(book_plan: BookPlan, chunks: list[EvidenceChunk]) -> list[ChapterPlan]:
    chunk_map = {chunk.chunk_id: chunk for chunk in chunks}
    plans: list[ChapterPlan] = []
    for chapter in book_plan.chapters:
        points: list[KnowledgePoint] = []
        for section in chapter.sections:
            for point_index, point_id in enumerate(section.knowledge_point_ids, start=len(points) + 1):
                point_chunks = [chunk_id for chunk_id in section.primary_material_ids if chunk_id in chunk_map]
                title = point_id
                points.append(
                    KnowledgePoint(
                        knowledge_point_id=f"kp_{chapter.chapter_no:02d}_{point_index:02d}",
                        title=title,
                        chunk_ids=point_chunks,
                        summary=_summary_for_chunks(point_chunks, chunk_map),
                        order_index=point_index,
                        difficulty_level=_difficulty_level(title),
                        prerequisite_ids=[],
                        cluster_id=_cluster_id(title),
                    )
                )
        _attach_simple_prerequisites(points)
        plans.append(
            ChapterPlan(
                chapter_id=chapter.chapter_id,
                title=chapter.title,
                learning_goals=chapter.learning_goals,
                knowledge_points=points,
                evidence_chunk_ids=chapter.primary_material_ids,
                activities=[
                    "阅读章节正文，梳理本章核心概念和操作要点。",
                    "结合示范视频完成观察、判断和迁移分析。",
                ],
                learning_path=[point.knowledge_point_id for point in points],
            )
        )
    return plans


def review_book_plan(book_plan: BookPlan, chunks: list[EvidenceChunk]) -> list[ReviewIssue]:
    chunk_map = {chunk.chunk_id: chunk for chunk in chunks}
    issues: list[ReviewIssue] = []
    seen_primary: Counter[str] = Counter()
    for chapter in book_plan.chapters:
        if not chapter.title.strip():
            issues.append(ReviewIssue("high", chapter.chapter_id, "章节标题为空。", "在 XLSX 或自动规划中补齐章节标题。"))
        if not chapter.sections:
            issues.append(ReviewIssue("high", chapter.chapter_id, "章节没有节次。", "至少保留一个节次和知识点。"))
        if not chapter.primary_material_ids:
            issues.append(ReviewIssue("high", chapter.chapter_id, "章节没有主素材。", "为该章分配正文生成主素材。"))
        for material_id in chapter.primary_material_ids:
            seen_primary[material_id] += 1
        has_video = any((chunk_map.get(mid) and chunk_map[mid].source_type in {"video_segment", "video", "audio_segment"}) for mid in chapter.primary_material_ids)
        has_practice = any(_difficulty_level(point) == "practice" for section in chapter.sections for point in section.knowledge_point_ids)
        if has_practice and not has_video:
            issues.append(ReviewIssue("medium", chapter.chapter_id, "实操章节缺少主视频。", "补充示范视频或调整素材归属。"))
        if len(chapter.sections) > 8:
            issues.append(ReviewIssue("low", chapter.chapter_id, "章节节次数较多。", "考虑拆分章节或收缩知识点。"))
    duplicates = sorted(material_id for material_id, count in seen_primary.items() if count > 1)
    for material_id in duplicates[:20]:
        issues.append(ReviewIssue("medium", material_id, "素材被多个章节设置为主归属。", "只保留一个主归属，其余章节改为引用归属。"))
    return issues


def render_book_outline_markdown(book_plan: BookPlan) -> str:
    lines = [f"# {book_plan.title} 教材大纲", ""]
    for chapter in book_plan.chapters:
        lines.extend([f"## 第{chapter.chapter_no}章 {chapter.title}", ""])
        for section_index, section in enumerate(chapter.sections, start=1):
            lines.extend([f"### {chapter.chapter_no}.{section_index} {section.title}", ""])
            for point_index, point in enumerate(section.knowledge_point_ids, start=1):
                lines.append(f"- {chapter.chapter_no}.{section_index}.{point_index} {point}")
            lines.append("")
    return "\n".join(lines).strip() + "\n"


def render_book_plan_review_markdown(title: str, issues: list[ReviewIssue]) -> str:
    lines = [f"# {title} 全书规划审核", ""]
    if not issues:
        lines.append("未发现全书规划门禁问题。")
        return "\n".join(lines) + "\n"
    for issue in issues:
        lines.append(f"- [{issue.severity}] {issue.location}：{issue.message} 建议：{issue.suggestion}")
    return "\n".join(lines) + "\n"


def _build_sections(
    *,
    chapter_id: str,
    chapter_no: int,
    chunks: list[EvidenceChunk],
    manifest_by_id: dict[str, ManifestMaterial],
    max_knowledge_points: int,
) -> list[BookSectionPlan]:
    by_section: dict[str, list[EvidenceChunk]] = defaultdict(list)
    for chunk in chunks:
        manifest = _manifest_for_chunk(chunk, manifest_by_id)
        section_title = _first_text(
            manifest.section if manifest else "",
            manifest.knowledge_point if manifest else "",
            chunk.metadata.get("section"),
            chunk.metadata.get("section_title"),
            chunk.metadata.get("knowledge_point"),
            chunk.title,
            "学习内容",
        )
        by_section[section_title].append(chunk)

    sections: list[BookSectionPlan] = []
    for section_index, (section_title, section_chunks) in enumerate(
        sorted(by_section.items(), key=lambda item: (_section_order(item[0], item[1]), item[0])),
        start=1,
    ):
        point_titles = _dedupe(
            [
                _first_text(
                    (_manifest_for_chunk(chunk, manifest_by_id).knowledge_point if _manifest_for_chunk(chunk, manifest_by_id) else ""),
                    chunk.metadata.get("knowledge_point"),
                    chunk.title,
                )
                for chunk in section_chunks
            ]
        )[:max_knowledge_points]
        primary_ids = _budget_primary_materials(section_chunks)
        sections.append(
            BookSectionPlan(
                section_id=f"{chapter_id}_section_{section_index:02d}",
                section_no=f"{chapter_no}.{section_index}",
                title=section_title,
                knowledge_point_ids=point_titles,
                primary_material_ids=primary_ids,
                reference_material_ids=[chunk.chunk_id for chunk in section_chunks if chunk.chunk_id not in primary_ids],
                recommended_video_ids=[
                    chunk.chunk_id
                    for chunk in section_chunks
                    if chunk.source_type in {"video_segment", "video", "audio_segment"}
                ][:DEFAULT_CHAPTER_VIDEO_BUDGET],
            )
        )
    return sections


def _budget_primary_materials(chunks: list[EvidenceChunk]) -> list[str]:
    videos = [chunk.chunk_id for chunk in chunks if chunk.source_type in {"video_segment", "video", "audio_segment"}]
    documents = [chunk.chunk_id for chunk in chunks if chunk.source_type not in {"video_segment", "video", "audio_segment"}]
    return _dedupe(videos[:DEFAULT_CHAPTER_VIDEO_BUDGET] + documents[:DEFAULT_CHAPTER_DOCUMENT_BUDGET])


def _best_manifest_by_id(manifest_rows: list[ManifestMaterial]) -> dict[str, ManifestMaterial]:
    grouped: dict[str, list[ManifestMaterial]] = defaultdict(list)
    for row in manifest_rows:
        if row.material_id:
            grouped[row.material_id].append(row)
        if row.path:
            grouped[row.path].append(row)
            grouped[Path(row.path).name].append(row)

    result: dict[str, ManifestMaterial] = {}
    for key, rows in grouped.items():
        result[key] = sorted(rows, key=_manifest_priority)[0]
    return result


def _manifest_priority(row: ManifestMaterial) -> tuple[int, int, float, str]:
    inactive_penalty = 1 if not row.active_for_index else 0
    confirmation_penalty = 1 if row.needs_confirmation else 0
    return (inactive_penalty, confirmation_penalty, -row.confidence, row.title)


def _manifest_for_chunk(chunk: EvidenceChunk, manifest_by_id: dict[str, ManifestMaterial]) -> ManifestMaterial | None:
    candidates = [
        chunk.chunk_id,
        chunk.asset_id,
        chunk.metadata.get("asset_id", ""),
        chunk.metadata.get("source_asset_id", ""),
        chunk.metadata.get("source_video", ""),
        Path(str(chunk.metadata.get("source_video", ""))).name if chunk.metadata.get("source_video") else "",
        chunk.locator.path,
        Path(chunk.locator.path).name if chunk.locator.path else "",
        chunk.locator.original_path,
        Path(chunk.locator.original_path).name if chunk.locator.original_path else "",
    ]
    for candidate in candidates:
        if candidate and str(candidate) in manifest_by_id:
            return manifest_by_id[str(candidate)]
    return None


def _chapter_order(
    title: str,
    chunks: list[EvidenceChunk],
    manifest_rows: list[ManifestMaterial],
    curriculum_index: dict[str, int] | None = None,
) -> int:
    if curriculum_index and title in curriculum_index:
        return curriculum_index[title]
    for row in manifest_rows:
        if row.chapter == title:
            match = re.search(r"\d+", row.chapter_code or row.chapter)
            if match:
                return int(match.group())
    orders = [int(chunk.metadata.get("chapter_order")) for chunk in chunks if str(chunk.metadata.get("chapter_order", "")).isdigit()]
    return min(orders) if orders else 999


def _section_order(title: str, chunks: list[EvidenceChunk]) -> int:
    match = re.search(r"\d+", title)
    if match:
        return int(match.group())
    orders = [int(chunk.metadata.get("knowledge_order")) for chunk in chunks if str(chunk.metadata.get("knowledge_order", "")).isdigit()]
    return min(orders) if orders else 999


def _summary_for_chunks(chunk_ids: list[str], chunk_map: dict[str, EvidenceChunk]) -> str:
    for chunk_id in chunk_ids:
        chunk = chunk_map.get(chunk_id)
        if chunk and chunk.summary:
            return chunk.summary
    return ""


def _difficulty_level(title: str) -> str:
    if any(term in title for term in ("操作", "引弧", "送丝", "收弧", "焊接")):
        return "practice"
    if any(term in title for term in ("特点", "适用", "范围", "缺陷", "质量", "综合")):
        return "advanced"
    return "basic"


def _cluster_id(title: str) -> str:
    level = _difficulty_level(title)
    return {"basic": "concept", "practice": "operation", "advanced": "extension"}[level]


def _attach_simple_prerequisites(points: list[KnowledgePoint]) -> None:
    basic_ids = [point.knowledge_point_id for point in points if point.difficulty_level == "basic"]
    for point in points:
        if point.difficulty_level != "basic":
            point.prerequisite_ids = basic_ids[:1]


def _field(row: dict[str, str], *names: str) -> str:
    for name in names:
        value = row.get(_normalize_header(name), "")
        if value:
            return value
    return ""


def _parse_float(value: str, default: float) -> float:
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return default


def _parse_bool(value: str, *, default: bool) -> bool:
    text = str(value or "").strip().lower()
    if not text:
        return default
    return text in {"1", "true", "yes", "y", "是", "启用", "有效", "active"}


def _find_matching_chapter(default_title: str, counts: Counter[str]) -> str:
    if default_title in counts:
        return default_title
    normalized_default = _normalize_title(default_title)
    for title in counts:
        normalized_title = _normalize_title(title)
        if normalized_default and (normalized_default in normalized_title or normalized_title in normalized_default):
            return title
    return ""


def _code_sort_key(code: str) -> tuple[int, str]:
    text = str(code or "").strip()
    match = re.search(r"\d+", text)
    if match:
        return (int(match.group()), text)
    return (999, text)


def _normalize_title(value: str) -> str:
    return re.sub(r"[\s\-_/：:，,。、《》()（）]+", "", value.strip().lower())


def _normalize_header(value: str) -> str:
    return re.sub(r"[\s_：:()（）-]+", "", value.strip().lower())


def _first_text(*values: Any) -> str:
    for value in values:
        text = str(value or "").strip()
        if text and text.lower() != "nan":
            return text
    return ""


def _dedupe(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if value and value not in seen:
            seen.add(value)
            result.append(value)
    return result


def _slugify(value: str) -> str:
    slug = re.sub(r"[^a-zA-Z0-9\u4e00-\u9fff]+", "-", value).strip("-").lower()
    return slug or "digital-textbook"
